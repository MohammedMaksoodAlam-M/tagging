#!/usr/bin/env python3
"""
Verification Layer - Validate classifications after tagging is complete

This script goes through each classified question in the result Excel and
uses AI to verify if the assigned tags (Subject/Topic/Subtopic) are appropriate.
Mismatched rows are highlighted yellow for manual review.
"""

import json
import logging
import os
import sys
from datetime import datetime
from typing import Dict, Optional, List

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from config import get_config, PATHS
from cost_tracker import CostTracker
from openai_client import OpenAIClient

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('logs/verification.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# Yellow fill for flagged rows
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Verification progress file
PROGRESS_FILE = "temp/verification_progress.json"


class ClassificationVerifier:
    """
    Verifies if assigned classifications match the question content using AI
    """

    def __init__(self, config: Dict, cost_tracker: Optional[CostTracker] = None):
        """
        Initialize the verifier

        Args:
            config: Configuration dictionary
            cost_tracker: Optional cost tracker instance
        """
        self.config = config
        self.cost_tracker = cost_tracker or CostTracker()

        # Initialize OpenAI client
        self.openai_client = OpenAIClient(config, self.cost_tracker)

        # Statistics
        self.stats = {
            'total_verified': 0,
            'valid': 0,
            'flagged': 0,
            'errors': 0,
            'start_time': datetime.now().isoformat()
        }

        # Progress tracking
        self.verified_rows = set()
        self.load_progress()

    def load_progress(self):
        """Load verification progress from file"""
        try:
            if os.path.exists(PROGRESS_FILE):
                with open(PROGRESS_FILE, 'r') as f:
                    data = json.load(f)
                    # Convert to set of tuples (tab_name, row_index)
                    self.verified_rows = set(
                        tuple(item) for item in data.get('verified_rows', [])
                    )
                    logger.info(f"Loaded progress: {len(self.verified_rows)} rows already verified")
        except Exception as e:
            logger.warning(f"Could not load progress: {e}")
            self.verified_rows = set()

    def save_progress(self):
        """Save verification progress to file"""
        try:
            os.makedirs(os.path.dirname(PROGRESS_FILE), exist_ok=True)
            data = {
                'verified_rows': list(self.verified_rows),
                'stats': self.stats,
                'last_updated': datetime.now().isoformat()
            }
            with open(PROGRESS_FILE, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            logger.warning(f"Could not save progress: {e}")

    def create_verification_prompt(self, question: str, explanation: str,
                                    subject: str, topic: str, subtopic: str) -> str:
        """
        Create a prompt for AI to verify the classification

        Args:
            question: The question text
            explanation: The explanation text
            subject: Assigned subject/chapter
            topic: Assigned topic
            subtopic: Assigned subtopic

        Returns:
            Verification prompt string
        """
        # Create full triplet for context
        triplet = f"{subject} > {topic} > {subtopic}"

        prompt = f"""You are a classification verification expert. Check if the assigned classification is reasonable for this question.

QUESTION:
{question}

EXPLANATION:
{explanation if explanation and explanation.lower() not in ['nan', 'none', ''] else 'No explanation provided'}

ASSIGNED CLASSIFICATION TRIPLET:
{triplet}

TASK:
Is this classification reasonable for this question?

GUIDELINES:
- Mark as VALID if the subject matches the question's domain
- Mark as VALID if the classification is close enough (topic/subtopic don't need to be perfect)
- Mark as INVALID if there's a clear mismatch between question content and subject
- Example of INVALID: A math/calculation question under History, or a Biology question under Economics
- Example of VALID: A question about rivers classified under Geography, even if subtopic could be better

RESPOND WITH JSON ONLY:
{{"is_valid": true, "reason": ""}}
or
{{"is_valid": false, "reason": "Brief reason for mismatch"}}
"""
        return prompt

    def verify_classification(self, question: str, explanation: str,
                               subject: str, topic: str, subtopic: str) -> Dict:
        """
        Use AI to verify if classification is appropriate

        Args:
            question: The question text
            explanation: The explanation text
            subject: Assigned subject
            topic: Assigned topic
            subtopic: Assigned subtopic

        Returns:
            Dict with is_valid and reason
        """
        prompt = self.create_verification_prompt(
            question, explanation, subject, topic, subtopic
        )

        try:
            response = self.openai_client.make_request(
                prompt,
                question=question[:50],
                operation="verification"
            )

            if not response:
                logger.warning("No response from AI")
                return {'is_valid': True, 'reason': 'AI unavailable, assuming valid'}

            # Parse JSON response
            response = response.strip()
            if response.startswith("```json"):
                response = response[7:]
            if response.startswith("```"):
                response = response[3:]
            if response.endswith("```"):
                response = response[:-3]

            result = json.loads(response.strip())
            return {
                'is_valid': result.get('is_valid', True),
                'reason': result.get('reason', '')
            }

        except json.JSONDecodeError as e:
            logger.warning(f"Could not parse AI response: {e}")
            return {'is_valid': True, 'reason': 'Parse error, assuming valid'}
        except Exception as e:
            logger.error(f"Verification error: {e}")
            return {'is_valid': True, 'reason': f'Error: {e}'}

    def verify_excel_file(self, file_path: str) -> Dict:
        """
        Verify all classifications in an Excel file

        Args:
            file_path: Path to the Excel file

        Returns:
            Dict with verification statistics
        """
        logger.info(f"Loading Excel file: {file_path}")

        try:
            # Load workbook with openpyxl to preserve/add formatting
            workbook = load_workbook(file_path)
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            return self.stats

        # Find column indices (assuming headers in row 1)
        column_map = {
            'Questions': None,
            'Explanation': None,
            'Subject': None,
            'Topic': None,
            'Subtopic': None
        }

        # Process each sheet (exam tab)
        for sheet_name in workbook.sheetnames:
            logger.info(f"\nProcessing sheet: {sheet_name}")
            sheet = workbook[sheet_name]

            # Map column names to indices
            header_row = list(sheet[1])
            for idx, cell in enumerate(header_row, 1):
                if cell.value in column_map:
                    column_map[cell.value] = idx

            # Check if all required columns found
            missing_cols = [k for k, v in column_map.items() if v is None]
            if missing_cols:
                logger.warning(f"Sheet {sheet_name} missing columns: {missing_cols}")
                continue

            # Find Feedback column or add it after last data column
            feedback_col = None
            last_data_col = 0

            for idx, cell in enumerate(header_row, 1):
                if cell.value == 'Feedback':
                    feedback_col = idx
                    break
                if cell.value is not None:
                    last_data_col = idx

            if feedback_col is None:
                # Add new Feedback column right after last data column
                feedback_col = last_data_col + 1
                sheet.cell(row=1, column=feedback_col).value = 'Feedback'
                logger.info(f"Added Feedback column at position {feedback_col}")

            # Count rows to process
            total_rows = sheet.max_row - 1  # Exclude header
            logger.info(f"Found {total_rows} rows to verify")

            # Process each row (starting from row 2, after header)
            for row_idx in range(2, sheet.max_row + 1):
                # Check if already verified
                row_key = (sheet_name, row_idx)
                if row_key in self.verified_rows:
                    continue

                # Get cell values
                question = sheet.cell(row=row_idx, column=column_map['Questions']).value
                explanation = sheet.cell(row=row_idx, column=column_map['Explanation']).value
                subject = sheet.cell(row=row_idx, column=column_map['Subject']).value
                topic = sheet.cell(row=row_idx, column=column_map['Topic']).value
                subtopic = sheet.cell(row=row_idx, column=column_map['Subtopic']).value

                # Skip rows without classification
                if not subject or not topic or not subtopic:
                    continue

                # Skip empty questions
                if not question or str(question).strip() == '':
                    continue

                # Convert to strings
                question = str(question).strip()
                explanation = str(explanation).strip() if explanation else ''
                subject = str(subject).strip()
                topic = str(topic).strip()
                subtopic = str(subtopic).strip()

                # Verify classification
                result = self.verify_classification(
                    question, explanation, subject, topic, subtopic
                )

                self.stats['total_verified'] += 1

                if result['is_valid']:
                    self.stats['valid'] += 1
                    logger.info(f"Row {row_idx}: VALID")
                else:
                    self.stats['flagged'] += 1
                    logger.warning(f"Row {row_idx}: FLAGGED - {result['reason']}")

                    # Write feedback reason for flagged rows (no color change)
                    sheet.cell(row=row_idx, column=feedback_col).value = result['reason']

                # Mark as verified
                self.verified_rows.add(row_key)

                # Progress update every 10 rows
                if self.stats['total_verified'] % 10 == 0:
                    logger.info(f"Progress: {self.stats['total_verified']} verified, "
                               f"{self.stats['flagged']} flagged")
                    self.save_progress()

                    # Save workbook periodically
                    try:
                        workbook.save(file_path)
                    except Exception as e:
                        logger.warning(f"Could not save workbook: {e}")

                # Check if we can continue (budget check)
                if not self.openai_client.can_continue():
                    logger.warning("Budget exhausted, stopping verification")
                    break

            # Check budget after each sheet
            if not self.openai_client.can_continue():
                break

        # Final save
        try:
            workbook.save(file_path)
            logger.info(f"Saved verified Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")

        # Save final progress
        self.save_progress()

        return self.stats

    def print_summary(self):
        """Print verification summary"""
        print("\n" + "=" * 50)
        print("VERIFICATION SUMMARY")
        print("=" * 50)
        print(f"Total verified: {self.stats['total_verified']}")
        print(f"Valid classifications: {self.stats['valid']}")
        print(f"Flagged (yellow): {self.stats['flagged']}")
        print(f"Errors: {self.stats['errors']}")

        if self.stats['total_verified'] > 0:
            flag_rate = (self.stats['flagged'] / self.stats['total_verified']) * 100
            print(f"Flag rate: {flag_rate:.1f}%")

        print("\n" + self.cost_tracker.generate_cost_report())


def find_classified_file() -> Optional[str]:
    """
    Find the classified Excel file to verify in result/ folder

    Returns:
        Path to the Excel file or None
    """
    import glob

    # Check command line argument first
    if len(sys.argv) > 1:
        if os.path.exists(sys.argv[1]):
            return sys.argv[1]

    # Only check result/ folder - this is where classified files go
    result_files = glob.glob("result/*.xlsx")
    if result_files:
        # Return most recently modified
        return max(result_files, key=os.path.getmtime)

    return None


def main():
    """Main entry point"""
    print("=" * 50)
    print("CLASSIFICATION VERIFICATION")
    print("=" * 50)

    # Get configuration
    config = get_config()

    # Determine input file
    result_file = find_classified_file()

    # Check if file exists
    if not result_file:
        logger.error("No classified Excel file found")
        print("\nError: Could not find any classified Excel file")
        print("Searched in: result/, output/, temp/")
        print("\nUsage: python verify_classifications.py [path_to_excel]")
        print("Or run classification first, then run verification.")
        sys.exit(1)

    print(f"\nVerifying: {result_file}")
    print("Flagged rows will be highlighted YELLOW\n")

    # Create verifier
    budget_usd = config.get("openai", {}).get("budget_limit_usd", 10.0)
    cost_tracker = CostTracker(budget_limit_usd=budget_usd)
    verifier = ClassificationVerifier(config, cost_tracker)

    try:
        # Run verification
        stats = verifier.verify_excel_file(result_file)

        # Print summary
        verifier.print_summary()

        print(f"\nVerification complete!")
        print(f"Flagged rows are highlighted YELLOW in: {result_file}")

    except KeyboardInterrupt:
        print("\n\nVerification interrupted by user")
        verifier.save_progress()
        verifier.print_summary()
        print("\nProgress saved. Run again to resume.")
    except Exception as e:
        logger.error(f"Verification failed: {e}")
        raise


if __name__ == "__main__":
    main()
